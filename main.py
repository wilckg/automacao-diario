import time
import re
import unicodedata
from datetime import datetime
from getpass import getpass

from openpyxl import load_workbook

from playwright.sync_api import (
    sync_playwright,
    TimeoutError as PlaywrightTimeoutError
)


DIARIO_URL = "https://diariofic.sp.senai.br/"

def normalizar_nome(nome):
    if nome is None:
        return ""

    nome = str(nome).strip().upper()

    nome = unicodedata.normalize(
        "NFKD",
        nome
    )

    nome = "".join(
        caractere
        for caractere in nome
        if not unicodedata.combining(caractere)
    )

    return " ".join(
        nome.split()
    )

def converter_hora_excel(valor):
    if valor is None:
        raise ValueError(
            "A duração da aula está vazia."
        )

    if isinstance(valor, datetime):
        return valor.strftime("%H:%M")

    if hasattr(valor, "strftime"):
        return valor.strftime("%H:%M")

    valor = str(valor).strip()

    if ":" in valor:
        partes = valor.split(":")

        hora = int(partes[0])
        minuto = int(partes[1])

        return f"{hora:02d}:{minuto:02d}"

    raise ValueError(
        f"Formato de hora inválido: {valor}"
    )

def obter_credenciais():
    print("\n=== Diário Eletrônico SENAI ===\n")

    nif = input("Digite seu NIF: ").strip()
    senha = getpass("Digite sua senha: ")

    if not nif:
        raise ValueError("O NIF não pode estar vazio.")

    if not senha:
        raise ValueError("A senha não pode estar vazia.")

    return nif, senha


def login_realizado(page):
    try:
        # Primeiro valida se chegou na página /home
        page.wait_for_url(
            "**/home",
            timeout=6000
        )

        # Depois valida um elemento que só existe após o login
        page.get_by_text(
            "Usuário:",
            exact=False
        ).wait_for(
            state="visible",
            timeout=5000
        )

        return True

    except PlaywrightTimeoutError:
        return False


def realizar_login(page, nif, senha, max_tentativas=3):
    print("\nAbrindo Diário Eletrônico...")

    for tentativa in range(1, max_tentativas + 1):

        print(
            f"\nTentativa de login "
            f"{tentativa}/{max_tentativas}"
        )

        # Volta para a página de login a cada tentativa
        page.goto(
            DIARIO_URL,
            wait_until="domcontentloaded",
            timeout=30000
        )

        campo_nif = page.locator(
            'input[name="aIdentificacao"]'
        )

        campo_senha = page.locator(
            'input[name="aSenha"]'
        )

        # Aguarda os campos realmente aparecerem
        campo_nif.wait_for(
            state="visible",
            timeout=10000
        )

        campo_senha.wait_for(
            state="visible",
            timeout=10000
        )

        # Limpa antes de preencher
        campo_nif.fill("")
        campo_senha.fill("")

        campo_nif.fill(nif)
        campo_senha.fill(senha)

        # Validação temporária
        print(
            "NIF preenchido:",
            campo_nif.input_value()
        )

        page.get_by_role(
            "button",
            name="Entrar"
        ).click()

        print("Aguardando resposta do sistema...")

        if login_realizado(page):
            print("\nLogin realizado com sucesso.")
            return True

        print(
            "Não foi possível confirmar o login."
        )

        # Podemos verificar se existe alguma mensagem
        # contendo "incorret"
        mensagem_erro = page.get_by_text(
            "incorret",
            exact=False
        )

        if mensagem_erro.count() > 0:
            try:
                if mensagem_erro.first.is_visible():
                    print(
                        "O sistema informou NIF ou senha incorretos."
                    )
            except Exception:
                pass

        if tentativa < max_tentativas:
            espera = tentativa * 2

            print(
                f"Tentando novamente em "
                f"{espera} segundos..."
            )

            time.sleep(espera)

    print(
        "\nNão foi possível realizar o login "
        "após todas as tentativas."
    )

    return False


def selecionar_unidade_e_abrir_turmas(page):
    print("\nCarregando unidades...")

    select = page.locator("select").first

    select.wait_for(
        state="visible",
        timeout=10000
    )

    opcoes = select.locator("option")

    unidades = []

    for i in range(opcoes.count()):

        opcao = opcoes.nth(i)

        nome = (
            opcao.text_content()
            or ""
        ).strip()

        valor = opcao.get_attribute(
            "value"
        )

        if valor and nome:
            unidades.append({
                "nome": nome,
                "valor": valor
            })

    if not unidades:
        raise RuntimeError(
            "Nenhuma unidade encontrada."
        )

    print("\n=== UNIDADES ===\n")

    for i, unidade in enumerate(
        unidades,
        start=1
    ):
        print(
            f"{i}. {unidade['nome']}"
        )

    while True:

        try:
            escolha = int(
                input(
                    "\nEscolha a unidade: "
                )
            )

            if 1 <= escolha <= len(unidades):
                break

            print(
                "Escolha uma opção válida."
            )

        except ValueError:

            print(
                "Digite apenas o número da unidade."
            )

    unidade = unidades[
        escolha - 1
    ]

    print(
        f"\nSelecionando: {unidade['nome']}"
    )

    select.select_option(
        value=unidade["valor"]
    )

    print("Unidade selecionada.")

    page.get_by_role(
        "link",
        name="Turmas"
    ).click()

    page.wait_for_load_state(
        "domcontentloaded"
    )

    print("Página de turmas aberta.")

def selecionar_turma(page):
    print("\nCarregando turmas...")

    page.wait_for_load_state("networkidle")

    tabela = page.locator("#tbDados")

    tabela.wait_for(
        state="visible",
        timeout=15000
    )

    linhas = tabela.locator("tbody tr")

    # linhas.first.wait_for(
    #     state="visible",
    #     timeout=15000
    # )

    quantidade_linhas = linhas.count()

    print(f"Turmas encontradas na tabela: {quantidade_linhas}")

    if quantidade_linhas == 0:
        raise RuntimeError(
            "Nenhuma turma encontrada na tabela."
        )

    turmas = []

    for i in range(quantidade_linhas):
        linha = linhas.nth(i)

        colunas = linha.locator("td")

        quantidade_colunas = colunas.count()

        if quantidade_colunas < 5:
            continue

        tipo_curso = (
            colunas.nth(0).inner_text()
            or ""
        ).strip()

        span_curso = colunas.nth(1).locator(
            "span[title]"
        )

        if span_curso.count() > 0:
            curso = (
                span_curso.first.get_attribute("title")
                or ""
            ).strip()
        else:
            curso = (
                colunas.nth(1).inner_text()
                or ""
            ).strip()

        codigo_turma = (
            colunas.nth(2).inner_text()
            or ""
        ).strip()

        periodo = (
            colunas.nth(3).inner_text()
            or ""
        ).strip()

        botao_editar = linha.locator(
            "i.fa-edit.edit"
        )

        if botao_editar.count() == 0:
            print(
                f"Aviso: turma {codigo_turma} "
                "não possui botão de editar."
            )
            continue

        id_turma = botao_editar.get_attribute(
            "data-id"
        )

        turmas.append({
            "tipo_curso": tipo_curso,
            "curso": curso,
            "codigo": codigo_turma,
            "periodo": periodo,
            "id": id_turma,
            "linha": linha
        })

    if not turmas:
        raise RuntimeError(
            "Nenhuma turma válida encontrada."
        )

    print("\n=== TURMAS DISPONÍVEIS ===\n")

    for i, turma in enumerate(
        turmas,
        start=1
    ):
        print(
            f"{i}. "
            f"{turma['codigo']} | "
            f"{turma['curso']} | "
            f"{turma['periodo']}"
        )

    while True:
        try:
            escolha = int(
                input(
                    "\nEscolha a turma: "
                )
            )

            if 1 <= escolha <= len(turmas):
                break

            print(
                "Escolha uma opção válida."
            )

        except ValueError:
            print(
                "Digite apenas o número da turma."
            )

    turma_selecionada = turmas[
        escolha - 1
    ]

    print(
        "\nTurma selecionada:"
    )

    print(
        f"Código: {turma_selecionada['codigo']}"
    )

    print(
        f"Curso: {turma_selecionada['curso']}"
    )

    print(
        f"Período: {turma_selecionada['periodo']}"
    )

    print(
        f"ID interno: {turma_selecionada['id']}"
    )

    return turma_selecionada

def abrir_turma_para_edicao(page, turma):
    print(
        f"\nAbrindo turma "
        f"{turma['codigo']}..."
    )

    botao_editar = turma["linha"].locator(
        "i.fa-edit.edit"
    )

    botao_editar.click()

    page.wait_for_load_state(
        "domcontentloaded"
    )

    print(
        "Tela de edição da turma aberta."
    )

def abrir_registro_aula(page):
    print("\nAbrindo registro de aula...")

    botao = page.locator("#NovaAula")

    botao.wait_for(
        state="visible",
        timeout=10000
    )

    botao.click()

    print("Aguardando formulário de registro...")

    page.get_by_text(
        "Diário de Classe",
        exact=False
    ).wait_for(
        state="visible",
        timeout=10000
    )

    print("Formulário de registro de aula aberto.")

def ler_alunos_diario(page):
    print("\n=== ALUNOS DO DIÁRIO ===")

    tabela = page.locator("#tbAlunos")

    tabela.wait_for(
        state="visible",
        timeout=10000
    )

    linhas = tabela.locator("tbody tr")

    quantidade = linhas.count()

    print(
        f"Quantidade de alunos encontrados: {quantidade}"
    )

    alunos = []

    for i in range(quantidade):
        linha = linhas.nth(i)

        colunas = linha.locator("td")

        if colunas.count() < 4:
            continue

        matricula = (
            colunas.nth(0).inner_text()
            or ""
        ).strip()

        nome = (
            colunas.nth(1).inner_text()
            or ""
        ).strip()

        situacao = (
            colunas.nth(2).inner_text()
            or ""
        ).strip()

        checkbox = linha.locator(
            'input[name="chkSelect"]'
        )

        if checkbox.count() == 0:
            print(
                f"Aviso: checkbox não encontrado para {nome}"
            )
            continue

        alunos.append({
            "matricula": matricula,
            "nome": nome,
            "nome_normalizado": normalizar_nome(nome),
            "situacao": situacao,
            "checkbox": checkbox,
            "linha": linha
        })

    for aluno in alunos:
        print(
            f"{aluno['matricula']} | "
            f"{aluno['nome']} | "
            f"{aluno['situacao']}"
        )

    return alunos

def diagnosticar_tabela_alunos(page):
    print("\n=== DIAGNÓSTICO - ALUNOS ===")

    tabelas = page.locator("table")

    print(
        f"Tabelas encontradas: {tabelas.count()}"
    )

    for i in range(tabelas.count()):
        tabela = tabelas.nth(i)

        print(
            f"\n--- TABELA {i} ---"
        )

        try:
            print(
                tabela.inner_text()
            )
        except Exception as erro:
            print(
                f"Não foi possível ler: {erro}"
            )

    print("\n=== FIM DO DIAGNÓSTICO ===")

def preencher_dados_aula(
    page,
    data,
    horas,
    conteudo
):
    print("\n=== PREENCHENDO AULA ===")

    campo_data = page.locator(
        "#DtaAula"
    )

    campo_horas = page.locator(
        "#DtaHora"
    )

    campo_conteudo = page.locator(
        "#Conteudo"
    )

    campo_data.fill(data)
    campo_horas.fill(horas)
    campo_conteudo.fill(conteudo)

    print(
        f"Data: {campo_data.input_value()}"
    )

    print(
        f"Horas: {campo_horas.input_value()}"
    )

    print(
        f"Conteúdo: {campo_conteudo.input_value()}"
    )

def diagnosticar_presencas(alunos):
    print("\n=== ESTADO DOS CHECKBOXES ===")

    for aluno in alunos:
        marcado = aluno["checkbox"].is_checked()

        status = (
            "PRESENTE"
            if marcado
            else "FALTA"
        )

        print(
            f"{aluno['nome']} -> {status}"
        )

def definir_presenca(page, aluno, presente):
    """Define o estado de presença usando o toggle visual da página."""
    checkbox = aluno["checkbox"]
    estado_atual = checkbox.is_checked()

    if estado_atual == presente:
        return

    linha = aluno["linha"]
    coluna_presenca = linha.locator("td").nth(3)
    toggle = coluna_presenca.locator(".toggle-switch label")
    toggle.click()
    page.wait_for_timeout(150)

    estado_final = checkbox.is_checked()
    if estado_final != presente:
        raise RuntimeError(
            f"Não foi possível alterar a presença de {aluno['nome']}."
        )


def preencher_campo_oculto(locator, valor, eventos=("input", "change")):
    """Preenche elemento presente no DOM mesmo com ancestral display:none."""
    locator.evaluate(
        """
        (el, args) => {
            el.value = args.valor;
            for (const evento of args.eventos) {
                el.dispatchEvent(new Event(evento, { bubbles: true }));
            }
        }
        """,
        {"valor": valor, "eventos": list(eventos)}
    )


def registrar_falta(page, aluno, horas_falta, detalhamento="1"):
    """
    Registra falta para um aluno.

    detalhamento:
      1 = Não compensado
      2 = Dispensado
      3 = Compensação
      4 = Tolerância por atraso
    """
    definir_presenca(page=page, aluno=aluno, presente=False)

    linha = aluno["linha"]
    select_detalhamento = linha.locator("select.slcDtl")
    hora_falta = linha.locator("input.dtaHora")

    if select_detalhamento.count() == 0:
        raise RuntimeError(
            f"Detalhamento da falta não encontrado para {aluno['nome']}."
        )

    if hora_falta.count() == 0:
        raise RuntimeError(
            f"Campo de período da falta não encontrado para {aluno['nome']}."
        )

    preencher_campo_oculto(
        select_detalhamento,
        detalhamento,
        eventos=("change",)
    )

    preencher_campo_oculto(
        hora_falta,
        horas_falta,
        eventos=("input", "change")
    )

    presente_final = aluno["checkbox"].is_checked()
    detalhamento_final = select_detalhamento.input_value()
    horas_final = hora_falta.input_value()

    if presente_final:
        raise RuntimeError(
            f"{aluno['nome']} continuou marcado como presente."
        )

    if detalhamento_final != detalhamento:
        raise RuntimeError(
            f"Falha ao definir detalhamento da falta de {aluno['nome']}. "
            f"Esperado={detalhamento!r}, obtido={detalhamento_final!r}"
        )

    if horas_final != horas_falta:
        raise RuntimeError(
            f"Falha ao definir período da falta de {aluno['nome']}. "
            f"Esperado={horas_falta!r}, obtido={horas_final!r}"
        )

    print(
        f"{aluno['nome']} -> FALTA | "
        f"detalhamento={detalhamento_final} | período={horas_final}"
    )


def registrar_presenca(page, aluno):
    """Garante que o aluno esteja marcado como presente."""
    definir_presenca(page=page, aluno=aluno, presente=True)

    if not aluno["checkbox"].is_checked():
        raise RuntimeError(
            f"Não foi possível marcar presença para {aluno['nome']}."
        )

    print(f"{aluno['nome']} -> PRESENTE")


def localizar_aluno_por_nome(alunos, nome):
    nome_procurado = normalizar_nome(nome)
    for aluno in alunos:
        if aluno["nome_normalizado"] == nome_procurado:
            return aluno
    return None


def testar_falta(page, alunos, nome, horas_falta="03:00"):
    """Teste controlado de uma falta, sem salvar a aula."""
    aluno = localizar_aluno_por_nome(alunos, nome)

    if aluno is None:
        print(f"Aluno não encontrado: {nome}")
        return False

    print(f"\nAluno encontrado: {aluno['nome']}")
    print("Antes:", aluno["checkbox"].is_checked())

    registrar_falta(
        page=page,
        aluno=aluno,
        horas_falta=horas_falta,
        detalhamento="1"
    )

    linha = aluno["linha"]
    select_detalhamento = linha.locator("select.slcDtl")
    hora_falta = linha.locator("input.dtaHora")
    coluna_detalhamento = linha.locator("td").nth(4)

    print("Depois:", aluno["checkbox"].is_checked())
    print("Detalhamento:", select_detalhamento.input_value())
    print("Período:", hora_falta.input_value())
    print(
        "Coluna de ausência:",
        coluna_detalhamento.evaluate(
            """
            el => ({
                display: getComputedStyle(el).display,
                visibility: getComputedStyle(el).visibility
            })
            """
        )
    )

    return True

def interpretar_frequencia(valor):
    """
    Converte o símbolo da planilha em um status.

    . = PRESENTE
    I = FALTA
    """
    if valor is None:
        return "SEM_REGISTRO"

    valor = str(valor).strip().upper()

    if valor == ".":
        return "PRESENTE"

    if valor == "I":
        return "FALTA"

    return "DESCONHECIDO"


def encontrar_coluna_dia(ws, dia):
    """
    Procura o dia na linha 2 da planilha.

    Exemplo:
    dia=7 encontra a coluna cujo cabeçalho é '07'.
    """
    dia_procurado = str(dia).zfill(2)

    for coluna in range(1, ws.max_column + 1):
        valor = ws.cell(
            row=2,
            column=coluna
        ).value

        if valor is None:
            continue

        valor = str(valor).strip().zfill(2)

        if valor == dia_procurado:
            return coluna

    raise RuntimeError(
        f"Dia {dia_procurado} não encontrado "
        "na planilha de presença."
    )


def ler_frequencia_excel(
    caminho,
    dia,
    nome_aba="Lista de Presença"
):
    print(
        f"\n=== LENDO FREQUÊNCIA DO DIA "
        f"{str(dia).zfill(2)} ==="
    )

    workbook = load_workbook(
        caminho,
        data_only=True
    )

    if nome_aba not in workbook.sheetnames:
        raise RuntimeError(
            f"Aba '{nome_aba}' não encontrada. "
            f"Abas disponíveis: {workbook.sheetnames}"
        )

    ws = workbook[nome_aba]

    coluna_dia = encontrar_coluna_dia(
        ws,
        dia
    )

    print(
        f"Dia encontrado na coluna {coluna_dia}."
    )

    alunos = []

    # No seu modelo:
    # coluna B = Nome
    # linha 3 em diante = alunos
    for linha in range(3, ws.max_row + 1):
        nome = ws.cell(
            row=linha,
            column=2
        ).value

        if nome is None:
            continue

        nome = str(nome).strip()

        if not nome:
            continue

        simbolo = ws.cell(
            row=linha,
            column=coluna_dia
        ).value

        status = interpretar_frequencia(
            simbolo
        )

        alunos.append({
            "nome": nome,
            "nome_normalizado": normalizar_nome(nome),
            "simbolo": simbolo,
            "status": status,
            "linha_excel": linha
        })

    print(
        f"Alunos encontrados no Excel: "
        f"{len(alunos)}"
    )

    return alunos


def comparar_frequencia(
    alunos_diario,
    alunos_excel
):
    print(
        "\n=== COMPARAÇÃO EXCEL X DIÁRIO ==="
    )

    diario_por_nome = {
        aluno["nome_normalizado"]: aluno
        for aluno in alunos_diario
    }

    excel_por_nome = {
        aluno["nome_normalizado"]: aluno
        for aluno in alunos_excel
    }

    encontrados = []
    somente_excel = []
    somente_diario = []

    presentes = []
    faltas = []
    sem_registro = []

    for aluno_excel in alunos_excel:
        nome_norm = aluno_excel[
            "nome_normalizado"
        ]

        aluno_diario = diario_por_nome.get(
            nome_norm
        )

        if aluno_diario is None:
            somente_excel.append(
                aluno_excel
            )
            continue

        registro = {
            "excel": aluno_excel,
            "diario": aluno_diario
        }

        encontrados.append(
            registro
        )

        status = aluno_excel[
            "status"
        ]

        if status == "PRESENTE":
            presentes.append(
                registro
            )

        elif status == "FALTA":
            faltas.append(
                registro
            )

        else:
            sem_registro.append(
                registro
            )

    for aluno_diario in alunos_diario:
        if (
            aluno_diario["nome_normalizado"]
            not in excel_por_nome
        ):
            somente_diario.append(
                aluno_diario
            )

    print(
        f"\nEncontrados nos dois: "
        f"{len(encontrados)}"
    )

    print(
        f"Presentes: "
        f"{len(presentes)}"
    )

    print(
        f"Faltas: "
        f"{len(faltas)}"
    )

    print(
        f"Sem registro/desconhecido: "
        f"{len(sem_registro)}"
    )

    print(
        f"Somente no Excel: "
        f"{len(somente_excel)}"
    )

    print(
        f"Somente no Diário: "
        f"{len(somente_diario)}"
    )

    print(
        "\n=== PRESENTES ==="
    )

    for item in presentes:
        print(
            f"OK  "
            f"{item['diario']['nome']}"
        )

    print(
        "\n=== FALTAS ==="
    )

    for item in faltas:
        print(
            f"FALTA  "
            f"{item['diario']['nome']}"
        )

    if somente_excel:
        print(
            "\n=== NO EXCEL, MAS NÃO NO DIÁRIO ==="
        )

        for aluno in somente_excel:
            print(
                f"- {aluno['nome']}"
            )

    if somente_diario:
        print(
            "\n=== NO DIÁRIO, MAS NÃO NO EXCEL ==="
        )

        for aluno in somente_diario:
            print(
                f"- {aluno['nome']}"
            )

    if sem_registro:
        print(
            "\n=== SEM REGISTRO VÁLIDO ==="
        )

        for item in sem_registro:
            print(
                f"- {item['diario']['nome']} "
                f"| símbolo="
                f"{item['excel']['simbolo']!r}"
            )

    return {
        "encontrados": encontrados,
        "presentes": presentes,
        "faltas": faltas,
        "sem_registro": sem_registro,
        "somente_excel": somente_excel,
        "somente_diario": somente_diario
    }

def aplicar_frequencia(
    page,
    comparacao,
    horas_aula
):
    print(
        "\n=== APLICANDO FREQUÊNCIA ==="
    )

    for item in comparacao["presentes"]:
        aluno = item["diario"]

        registrar_presenca(
            page=page,
            aluno=aluno
        )

    for item in comparacao["faltas"]:
        aluno = item["diario"]

        registrar_falta(
            page=page,
            aluno=aluno,
            horas_falta=horas_aula,
            detalhamento="1"
        )

    print(
        "\nFrequência aplicada."
    )

def validar_comparacao_antes_de_aplicar(
    comparacao
):
    problemas = False

    if comparacao["somente_diario"]:
        problemas = True

    if comparacao["sem_registro"]:
        problemas = True

    if problemas:
        print(
            "\nATENÇÃO:"
        )

        print(
            "Existem divergências entre "
            "Excel e Diário."
        )

        print(
            "A frequência NÃO será aplicada."
        )

        return False

    return True

def ler_aulas_excel(
    caminho,
    nome_aba="Aulas"
):
    print("\n=== LENDO AULAS ===")

    workbook = load_workbook(
        caminho,
        data_only=True
    )

    if nome_aba not in workbook.sheetnames:
        raise RuntimeError(
            f"Aba '{nome_aba}' não encontrada. "
            f"Abas disponíveis: "
            f"{workbook.sheetnames}"
        )

    ws = workbook[nome_aba]

    aulas = []

    for linha in range(
        2,
        ws.max_row + 1
    ):
        data = ws.cell(
            row=linha,
            column=1
        ).value

        horas = ws.cell(
            row=linha,
            column=2
        ).value

        conteudo = ws.cell(
            row=linha,
            column=3
        ).value

        registrar = ws.cell(
            row=linha,
            column=4
        ).value

        # ignora linhas vazias
        if data is None:
            continue

        registrar = (
            str(registrar or "SIM")
            .strip()
            .upper()
        )

        if registrar not in (
            "SIM",
            "S",
            "1"
        ):
            continue

        # DATA
        if isinstance(data, datetime):
            data_obj = data
        else:
            data_obj = datetime.strptime(
                str(data),
                "%d/%m/%Y"
            )

        data_iso = data_obj.strftime(
            "%Y-%m-%d"
        )

        dia = data_obj.day

        # HORAS
        horas_formatadas = (
            converter_hora_excel(
                horas
            )
        )

        if not conteudo:
            raise RuntimeError(
                f"Conteúdo vazio na linha "
                f"{linha} da aba Aulas."
            )

        aulas.append({
            "data": data_obj,
            "data_iso": data_iso,
            "dia": dia,
            "horas": horas_formatadas,
            "conteudo": str(
                conteudo
            ).strip(),
            "linha_excel": linha
        })

    print(
        f"Aulas encontradas: "
        f"{len(aulas)}"
    )

    for aula in aulas:
        print(
            f"{aula['data'].strftime('%d/%m/%Y')} "
            f"| {aula['horas']} "
            f"| {aula['conteudo']}"
        )

    return aulas


def selecionar_aula_excel(aulas):
    """
    Permite escolher uma das aulas habilitadas na aba Aulas.
    Neste estágio processamos apenas UMA aula por vez, sem salvar,
    para validar o fluxo antes de criar o processamento em lote.
    """
    if not aulas:
        raise RuntimeError(
            "Nenhuma aula marcada para registro foi encontrada na aba 'Aulas'."
        )

    print("\n=== AULAS DISPONÍVEIS PARA TESTE ===\n")

    for i, aula in enumerate(aulas, start=1):
        print(
            f"{i}. "
            f"{aula['data'].strftime('%d/%m/%Y')} | "
            f"{aula['horas']} | "
            f"{aula['conteudo']}"
        )

    while True:
        try:
            escolha = int(
                input("\nEscolha a aula que deseja preparar: ")
            )

            if 1 <= escolha <= len(aulas):
                return aulas[escolha - 1]

            print("Escolha uma opção válida.")

        except ValueError:
            print("Digite apenas o número da aula.")


def exibir_resumo_aula(aula, comparacao):
    print("\n==========================================")
    print("=== RESUMO DA AULA QUE SERÁ PREPARADA ===")
    print("==========================================")

    print(
        "Data:",
        aula["data"].strftime("%d/%m/%Y")
    )
    print("Horas:", aula["horas"])
    print("Conteúdo:", aula["conteudo"])

    print(
        f"Presentes: {len(comparacao['presentes'])}"
    )
    print(
        f"Faltas: {len(comparacao['faltas'])}"
    )

    if comparacao["faltas"]:
        print("\nAlunos com falta:")

        for item in comparacao["faltas"]:
            print(
                f"- {item['diario']['nome']}"
            )

    if comparacao["somente_excel"]:
        print(
            "\nAlunos que estão somente no Excel "
            "(serão ignorados):"
        )

        for aluno in comparacao["somente_excel"]:
            print(f"- {aluno['nome']}")

    print("==========================================\n")


def extrair_datas_aulas_registradas(page):
    """
    Lê as aulas já cadastradas na tela da turma e retorna um conjunto
    com as datas no formato YYYY-MM-DD.

    A função procura datas visíveis na tabela de aulas já registradas.
    """
    datas_registradas = set()

    tabelas = page.locator("table")

    padrao_data = re.compile(
        r"\b(\d{2})/(\d{2})/(\d{4})\b"
    )

    for i in range(tabelas.count()):
        tabela = tabelas.nth(i)

        try:
            texto = tabela.inner_text()
        except Exception:
            continue

        for dia, mes, ano in padrao_data.findall(texto):
            try:
                data_obj = datetime.strptime(
                    f"{dia}/{mes}/{ano}",
                    "%d/%m/%Y"
                )
                datas_registradas.add(
                    data_obj.strftime("%Y-%m-%d")
                )
            except ValueError:
                continue

    return datas_registradas


def filtrar_aulas_ja_registradas(page, aulas):
    """
    Compara a aba Aulas com as aulas já cadastradas no Diário.

    Retorna:
      - aulas_pendentes
      - aulas_ignoradas
    """
    datas_registradas = extrair_datas_aulas_registradas(
        page
    )

    print("\n=== VERIFICAÇÃO DE DUPLICIDADE ===")

    if datas_registradas:
        print(
            f"Datas já encontradas no Diário: "
            f"{len(datas_registradas)}"
        )
    else:
        print(
            "Nenhuma data já registrada foi identificada."
        )

    aulas_pendentes = []
    aulas_ignoradas = []

    for aula in aulas:
        if aula["data_iso"] in datas_registradas:
            aulas_ignoradas.append(aula)

            print(
                f"IGNORAR - "
                f"{aula['data'].strftime('%d/%m/%Y')} "
                f"já está registrada."
            )
        else:
            aulas_pendentes.append(aula)

            print(
                f"PROCESSAR - "
                f"{aula['data'].strftime('%d/%m/%Y')}"
            )

    print(
        f"\nPendentes: {len(aulas_pendentes)}"
    )

    print(
        f"Ignoradas por duplicidade: "
        f"{len(aulas_ignoradas)}"
    )

    return aulas_pendentes, aulas_ignoradas


def salvar_aula(page):
    print("\nSalvando aula...")

    botao_salvar = page.locator("#btnSalvar")

    botao_salvar.wait_for(
        state="visible",
        timeout=10000
    )

    botao_salvar.click()

    # Aguarda o formulário de registro desaparecer.
    try:
        page.locator("#DtaAula").wait_for(
            state="hidden",
            timeout=15000
        )
    except PlaywrightTimeoutError:
        raise RuntimeError(
            "O formulário continuou aberto após clicar em Salvar. "
            "Pode ter ocorrido algum erro de validação."
        )

    # Confirma que voltamos para a tela onde é possível registrar outra aula.
    page.locator("#NovaAula").wait_for(
        state="visible",
        timeout=15000
    )

    print("Aula salva com sucesso.")

def confirmar_processamento_lote(aulas):
    print("\n======================================")
    print("=== AULAS QUE SERÃO REGISTRADAS ===")
    print("======================================\n")

    for i, aula in enumerate(
        aulas,
        start=1
    ):
        print(
            f"{i}. "
            f"{aula['data'].strftime('%d/%m/%Y')} "
            f"| {aula['horas']} "
            f"| {aula['conteudo']}"
        )

    print(
        f"\nTotal de aulas: {len(aulas)}"
    )

    resposta = input(
        "\nDeseja registrar TODAS essas aulas? [S/N]: "
    ).strip().upper()

    return resposta in (
        "S",
        "SIM"
    )

def processar_aula(
    page,
    aula,
    caminho_excel
):
    print("\n")
    print("=" * 60)

    print(
        f"PROCESSANDO AULA "
        f"{aula['data'].strftime('%d/%m/%Y')}"
    )

    print("=" * 60)

    # Proteção extra: revalida a data imediatamente antes de processar.
    # Isso evita duplicidade caso a página tenha mudado desde o início do lote.
    datas_registradas = extrair_datas_aulas_registradas(
        page
    )

    if aula["data_iso"] in datas_registradas:
        print(
            f"A aula de "
            f"{aula['data'].strftime('%d/%m/%Y')} "
            f"já está registrada. Ignorando."
        )
        return "IGNORADA"

    # Estamos na tela da turma.
    abrir_registro_aula(
        page
    )

    # Como cada novo formulário possui novos Locators,
    # lemos os alunos novamente.
    alunos_diario = ler_alunos_diario(
        page
    )

    # Frequência correspondente à data atual.
    alunos_excel = ler_frequencia_excel(
        caminho=caminho_excel,
        dia=aula["dia"]
    )

    comparacao = comparar_frequencia(
        alunos_diario=alunos_diario,
        alunos_excel=alunos_excel
    )

    exibir_resumo_aula(
        aula=aula,
        comparacao=comparacao
    )

    if not validar_comparacao_antes_de_aplicar(
        comparacao
    ):
        raise RuntimeError(
            f"A frequência de "
            f"{aula['data'].strftime('%d/%m/%Y')} "
            f"possui divergências."
        )

    preencher_dados_aula(
        page=page,
        data=aula["data_iso"],
        horas=aula["horas"],
        conteudo=aula["conteudo"]
    )

    aplicar_frequencia(
        page=page,
        comparacao=comparacao,
        horas_aula=aula["horas"]
    )

    salvar_aula(
        page
    )

def processar_todas_as_aulas(
    page,
    aulas,
    caminho_excel
):
    total = len(aulas)

    sucesso = []
    erros = []

    for indice, aula in enumerate(
        aulas,
        start=1
    ):
        data_formatada = (
            aula["data"].strftime(
                "%d/%m/%Y"
            )
        )

        print(
            f"\n\n[{indice}/{total}] "
            f"Iniciando {data_formatada}"
        )

        try:
            resultado = processar_aula(
                page=page,
                aula=aula,
                caminho_excel=caminho_excel
            )

            if resultado == "IGNORADA":
                print(
                    f"IGNORADA - {data_formatada}"
                )
                continue

            sucesso.append(
                data_formatada
            )

        except Exception as erro:
            print(
                f"\nERRO na aula "
                f"{data_formatada}: {erro}"
            )

            erros.append({
                "data": data_formatada,
                "erro": str(erro)
            })

            # Mais seguro parar no primeiro erro.
            break

    print("\n")
    print("=" * 60)
    print("=== RESULTADO DO PROCESSAMENTO ===")
    print("=" * 60)

    print(
        f"Aulas salvas: {len(sucesso)}"
    )

    for data in sucesso:
        print(
            f"OK - {data}"
        )

    if erros:
        print(
            f"\nErros: {len(erros)}"
        )

        for item in erros:
            print(
                f"ERRO - {item['data']} "
                f"| {item['erro']}"
            )

    print("=" * 60)



# def diagnosticar_turmas(page):
#     print("\n===== DIAGNÓSTICO DA PÁGINA DE TURMAS =====")

#     print("URL atual:")
#     print(page.url)

#     print("\nTítulo:")
#     print(page.title())

#     print("\nFrames:")
#     for frame in page.frames:
#         print(" -", frame.url)


#     print("\nQuantidade de #tbDados:")
#     print(
#         page.locator("#tbDados").count()
#     )

#     print("\nQuantidade de tbody:")
#     print(
#         page.locator("#tbDados tbody").count()
#     )

#     print("\nQuantidade de tr:")
#     print(
#         page.locator("#tbDados tbody tr").count()
#     )

#     print("\nQuantidade de td:")
#     print(
#         page.locator("#tbDados tbody td").count()
#     )

#     print("\nQuantidade de ícones editar:")
#     print(
#         page.locator("#tbDados i.edit").count()
#     )

#     print("\n==========================================")

def main():
    try:
        ARQUIVO_PRESENCA = (
            "lista_presenca_com_frequencia.xlsx"
        )

        aulas = ler_aulas_excel(
            caminho=ARQUIVO_PRESENCA
        )

        if not aulas:
            print(
                "\nNenhuma aula marcada "
                "para registro."
            )
            return

        nif, senha = obter_credenciais()

        with sync_playwright() as p:

            browser = p.chromium.launch(
                headless=False,
                slow_mo=300
            )

            context = browser.new_context()

            page = context.new_page()

            sucesso_login = realizar_login(
                page=page,
                nif=nif,
                senha=senha,
                max_tentativas=3
            )

            if not sucesso_login:
                print(
                    "\nEncerrando automação."
                )

                browser.close()
                return

            selecionar_unidade_e_abrir_turmas(
                page
            )

            turma = selecionar_turma(
                page
            )

            abrir_turma_para_edicao(
                page,
                turma
            )

            aulas_pendentes, aulas_ignoradas = (
                filtrar_aulas_ja_registradas(
                    page=page,
                    aulas=aulas
                )
            )

            if aulas_ignoradas:
                print(
                    "\nAs aulas já registradas foram "
                    "removidas automaticamente do lote."
                )

            if not aulas_pendentes:
                print(
                    "\nTodas as aulas marcadas no Excel "
                    "já estão registradas no Diário."
                )
                print(
                    "Nenhuma aula nova será cadastrada."
                )

                page.pause()
                browser.close()
                return

            if not confirmar_processamento_lote(
                aulas_pendentes
            ):
                print(
                    "\nProcessamento cancelado."
                )

                browser.close()
                return

            processar_todas_as_aulas(
                page=page,
                aulas=aulas_pendentes,
                caminho_excel=ARQUIVO_PRESENCA
            )

            page.pause()

            browser.close()

    except KeyboardInterrupt:
        print(
            "\n\nAutomação cancelada pelo usuário."
        )

    except FileNotFoundError as erro:
        print(
            "\nArquivo Excel não encontrado."
        )

        print(
            "Confirme se o arquivo está "
            "na mesma pasta do main.py."
        )

        print(
            f"Detalhes: {erro}"
        )

    except Exception as erro:
        print(
            f"\nErro: {erro}"
        )


if __name__ == "__main__":
    main()
